import os
import logging
import numpy as np
import pandas as pd
import pyreadstat
import openpyxl
from openpyxl.utils.cell import get_column_letter

logger = logging.getLogger(__name__)


def _normalize_inf_strings(df: pd.DataFrame) -> pd.DataFrame:
    """Convert common textual infinity markers to numeric infinities where possible."""
    if df is None or df.empty:
        return df

    df = df.copy()
    inf_map = {
        'inf': np.inf,
        '+inf': np.inf,
        'infinity': np.inf,
        '+infinity': np.inf,
        '-inf': -np.inf,
        '-infinity': -np.inf,
    }

    for col in df.columns:
        s = df[col]
        if pd.api.types.is_object_dtype(s) or pd.api.types.is_string_dtype(s):
            stripped = s.map(lambda x: x.strip().lower() if isinstance(x, str) else x)
            matched = stripped.isin(inf_map.keys()) if hasattr(stripped, 'isin') else pd.Series(False, index=s.index)
            if bool(matched.any()):
                logger.warning("Found textual infinity markers in column '%s': %s rows", col, int(matched.sum()))
                df.loc[matched, col] = stripped.loc[matched].map(inf_map)

                # If column now looks numeric, coerce it.
                coerced = pd.to_numeric(df[col], errors='ignore')
                df[col] = coerced

    return df


def read_dataframe(filename):
    """
    Reads a dataframe from a given file. Supports CSV, Excel (.xls/.xlsx), Pickle (.pkl), and SPSS (.sav).
    Applies a light post-processing pass to normalize textual infinity markers.
    """
    lower_name = filename.lower()
    logger.info("read_dataframe called for file: %s", filename)

    if lower_name.endswith('.csv'):
        logger.info("Detected CSV input")
        df = pd.read_csv(filename, na_values=['/'])

    elif lower_name.endswith(('.xls', '.xlsx')):
        logger.info("Detected Excel input")
        df = pd.read_excel(filename, engine='openpyxl', na_values=['/'])

    elif lower_name.endswith('.pkl'):
        logger.info("Detected Pickle input")
        df = pd.read_pickle(filename)

    elif lower_name.endswith('.sav'):
        logger.info("Detected SPSS input")
        # 1. user_missing=True converts SPSS user-defined missing codes to NaN
        df, meta = pyreadstat.read_sav(filename, user_missing=True)
        
        # 2. OPTIONAL but recommended: Convert coded categorical variables into their string labels
        # This allows core.characterize_columns to correctly classify them as 'multinomial_categorical'
        df = pyreadstat.set_value_labels(df, meta, formats_as_category=False)
        
        logger.info("SPSS metadata extracted: encoding=%s", getattr(meta, 'file_encoding', None))
        logger.debug("SPSS variable labels: %s", getattr(meta, 'variable_labels', None))
        logger.debug("SPSS value labels: %s", getattr(meta, 'value_labels', None))

    else:
        raise ValueError("Unsupported file format. Use .csv, .xls, .xlsx, .pkl, or .sav")

    logger.info("Raw dataframe loaded: shape=%s", df.shape)
    logger.info("Raw dtypes: %s", {col: str(dtype) for col, dtype in df.dtypes.items()})

    df = _normalize_inf_strings(df)

    logger.info("Post-processed dataframe ready: shape=%s", df.shape)
    logger.info("Post-processed dtypes: %s", {col: str(dtype) for col, dtype in df.dtypes.items()})
    return df


def extract_header_metadata(file_path, max_header_rows=5):
    """
    Scans an Excel file to resolve merged cells and extract background colors.
    Returns a structured dictionary of the top rows.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active

    merge_map = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        master_value = sheet.cell(row=min_row, column=min_col).value

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merge_map[(row, col)] = master_value

    header_matrix = []

    for row_idx in range(1, max_header_rows + 1):
        row_data = []

        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)

            if (row_idx, col_idx) in merge_map:
                val = merge_map[(row_idx, col_idx)]
            else:
                val = cell.value

            color_hex = None
            if cell.fill and cell.fill.start_color:
                color_hex = cell.fill.start_color.rgb
                if color_hex == '00000000' or type(color_hex) != str:
                    color_hex = None

            row_data.append({
                "col_index": col_idx,
                "column_letter": get_column_letter(col_idx),
                "value": str(val).strip() if val is not None else "",
                "raw_value": str(val) if val is not None else "",
                "color_hex": color_hex
            })

        header_matrix.append(row_data)

    return header_matrix


def resolve_visual_formatting(header_matrix):
    if not header_matrix:
        return header_matrix, []

    num_rows = len(header_matrix)
    num_cols = len(header_matrix[0])

    for row in header_matrix:
        current_section_header = None

        for cell in row:
            raw_val = cell.get("raw_value", "")
            stripped_val = cell.get("value", "")

            if stripped_val != "":
                if len(raw_val) - len(stripped_val) >= 4:
                    current_section_header = stripped_val
                else:
                    current_section_header = None
            elif current_section_header is not None:
                cell["value"] = current_section_header

    valid_pandas_columns = []

    for col_idx in range(num_cols):
        is_empty_divider = True
        for row_idx in range(num_rows):
            if header_matrix[row_idx][col_idx]["value"] != "":
                is_empty_divider = False
                break

        if not is_empty_divider:
            actual_excel_col = header_matrix[0][col_idx]["col_index"]
            valid_pandas_columns.append(actual_excel_col - 1)

    return header_matrix, valid_pandas_columns


def flatten_header_matrix(header_matrix, valid_cols):
    if not header_matrix:
        return []

    num_rows = len(header_matrix)
    clean_headers = []

    for col_idx in valid_cols:
        col_parts = []
        for row_idx in range(num_rows):
            val = header_matrix[row_idx][col_idx]["value"]
            if val == "":
                continue
            if not col_parts or col_parts[-1] != val:
                col_parts.append(val)

        if col_parts:
            final_name = "_".join(col_parts)
        else:
            final_name = f"Unnamed_Col_{col_idx}"

        clean_headers.append(final_name)

    return clean_headers


def load_clinical_excel_safely(file_path, header_rows_to_scan=5):
    raw_matrix = extract_header_metadata(file_path, max_header_rows=header_rows_to_scan)
    resolved_matrix, valid_cols = resolve_visual_formatting(raw_matrix)
    clean_headers = flatten_header_matrix(resolved_matrix, valid_cols)

    df = pd.read_excel(
        file_path,
        skiprows=header_rows_to_scan,
        usecols=valid_cols,
        names=clean_headers,
    )

    return _normalize_inf_strings(df)
